# Create new file: app/analytics/routes.py

from flask import render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required, current_user
from app.analytics import bp
from app.models import (User, WorkExtension, WorkExtensionStatus, Shift, ShiftStatus, 
                       Section, Unit, db, EmployeeType, UserRole)
from datetime import datetime, date, timedelta
from sqlalchemy import func, text, and_, or_
from collections import defaultdict
import calendar

def has_analytics_access():
    """Check if user has access to analytics (managers and above)"""
    return current_user.can_approve_leaves() or current_user.can_edit_schedule()

def get_user_scope():
    """Get the scope of data user can access"""
    if current_user.can_admin():
        return {
            'type': 'admin',
            'name': 'All Data',
            'department_id': None,
            'division_id': None,
            'section_id': None,
            'unit_id': None
        }
    
    # Check 4-level hierarchy access
    if hasattr(current_user, 'department_id') and current_user.department_id:
        return {
            'type': 'department',
            'name': f"{current_user.department.name} Department",
            'department_id': current_user.department_id,
            'division_id': None,
            'section_id': None,
            'unit_id': None
        }
    
    if hasattr(current_user, 'division_id') and current_user.division_id:
        return {
            'type': 'division',
            'name': f"{current_user.division.name} Division",
            'department_id': None,
            'division_id': current_user.division_id,
            'section_id': None,
            'unit_id': None
        }
    
    if current_user.section_id:
        return {
            'type': 'section',
            'name': f"{current_user.section.name} Section",
            'department_id': None,
            'division_id': None,
            'section_id': current_user.section_id,
            'unit_id': None
        }
    
    if current_user.unit_id:
        return {
            'type': 'unit',
            'name': f"{current_user.unit.name} Unit",
            'department_id': None,
            'division_id': None,
            'section_id': None,
            'unit_id': current_user.unit_id
        }
    
    return None

def filter_employees_by_scope(query, scope):
    """Filter employee query based on user's scope"""
    if scope['type'] == 'admin':
        return query
    elif scope['type'] == 'department':
        return query.filter(User.department_id == scope['department_id'])
    elif scope['type'] == 'division':
        return query.filter(User.division_id == scope['division_id'])
    elif scope['type'] == 'section':
        return query.filter(User.section_id == scope['section_id'])
    elif scope['type'] == 'unit':
        return query.filter(User.unit_id == scope['unit_id'])
    else:
        return query.filter(User.id == current_user.id)  # Fallback to own data only

@bp.route('/')
@login_required
def analytics_dashboard():
    """Analytics dashboard page"""
    if not has_analytics_access():
        flash('You do not have permission to access Analytics.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    scope = get_user_scope()
    if not scope:
        flash('Unable to determine your data access scope.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get date range options
    today = date.today()
    current_month_start = today.replace(day=1)
    last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    current_year_start = today.replace(month=1, day=1)
    
    date_ranges = {
        'current_month': {
            'name': 'Current Month',
            'start': current_month_start,
            'end': today
        },
        'last_month': {
            'name': 'Last Month',
            'start': last_month_start,
            'end': current_month_start - timedelta(days=1)
        },
        'current_year': {
            'name': 'Current Year',
            'start': current_year_start,
            'end': today
        },
        'last_30_days': {
            'name': 'Last 30 Days',
            'start': today - timedelta(days=30),
            'end': today
        }
    }
    
    # Get quick stats
    stats = get_quick_stats(scope)
    
    return render_template('analytics/dashboard.html', 
                         scope=scope,
                         date_ranges=date_ranges,
                         stats=stats)

@bp.route('/api/work-extension-data')
@login_required
def get_work_extension_data():
    """API endpoint for work extension analytics data"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    view_type = request.args.get('view_type', 'daily')  # daily, weekly, monthly
    group_by = request.args.get('group_by', 'total')   # total, section, unit, department, division
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    # Get work extension data
    data = get_work_extension_analytics(scope, start_date, end_date, view_type, group_by)
    
    return jsonify({
        'success': True,
        'data': data,
        'scope': scope,
        'parameters': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'view_type': view_type,
            'group_by': group_by
        }
    })

@bp.route('/api/overtime-data')
@login_required
def get_overtime_data():
    """API endpoint for overtime analytics data"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    view_type = request.args.get('view_type', 'daily')
    group_by = request.args.get('group_by', 'total')
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    # Get overtime data (calculated from shifts)
    data = get_overtime_analytics(scope, start_date, end_date, view_type, group_by)
    
    return jsonify({
        'success': True,
        'data': data,
        'scope': scope,
        'parameters': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'view_type': view_type,
            'group_by': group_by
        }
    })

def get_quick_stats(scope):
    """Get quick statistics for the dashboard"""
    today = date.today()
    current_month_start = today.replace(day=1)
    
    # Base query for employees in scope
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employee_ids = [emp.id for emp in employee_query.all()]
    
    if not employee_ids:
        return {
            'total_employees': 0,
            'current_month_extensions': 0,
            'current_month_extension_hours': 0,
            'pending_extensions': 0
        }
    
    # Work extension stats
    current_month_extensions = WorkExtension.query.filter(
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.extension_date >= current_month_start,
        WorkExtension.extension_date <= today,
        WorkExtension.status == WorkExtensionStatus.APPROVED
    ).count()
    
    current_month_extension_hours = db.session.query(
        func.sum(WorkExtension.extension_hours)
    ).filter(
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.extension_date >= current_month_start,
        WorkExtension.extension_date <= today,
        WorkExtension.status == WorkExtensionStatus.APPROVED
    ).scalar() or 0
    
    pending_extensions = WorkExtension.query.filter(
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.status == WorkExtensionStatus.PENDING
    ).count()
    
    return {
        'total_employees': len(employee_ids),
        'current_month_extensions': current_month_extensions,
        'current_month_extension_hours': float(current_month_extension_hours),
        'pending_extensions': pending_extensions
    }

def get_work_extension_analytics(scope, start_date, end_date, view_type, group_by):
    """Get work extension analytics data"""
    # Base query for employees in scope
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employee_ids = [emp.id for emp in employee_query.all()]
    
    if not employee_ids:
        return []
    
    # Query work extensions with EXPLICIT JOIN condition
    extensions_query = db.session.query(
        WorkExtension.extension_date,
        WorkExtension.extension_hours,
        User.section_id,
        User.unit_id
    ).join(User, WorkExtension.employee_id == User.id).filter(  # EXPLICIT JOIN
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.extension_date.between(start_date, end_date),
        WorkExtension.status == WorkExtensionStatus.APPROVED,
        WorkExtension.extension_hours.isnot(None)
    )
    
    # Add 4-level hierarchy fields if available
    if hasattr(User, 'department_id'):
        extensions_query = extensions_query.add_columns(User.department_id)
    if hasattr(User, 'division_id'):
        extensions_query = extensions_query.add_columns(User.division_id)
    
    extensions = extensions_query.all()
    
    # Process data based on view_type and group_by
    if view_type == 'daily':
        return process_daily_data(extensions, start_date, end_date, group_by, 'extension')
    elif view_type == 'weekly':
        return process_weekly_data(extensions, start_date, end_date, group_by, 'extension')
    elif view_type == 'monthly':
        return process_monthly_data(extensions, start_date, end_date, group_by, 'extension')
    else:
        return []

def get_overtime_analytics(scope, start_date, end_date, view_type, group_by):
    """Get overtime analytics data from shifts"""
    # Base query for employees in scope
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employee_ids = [emp.id for emp in employee_query.all()]
    
    if not employee_ids:
        return []
    
    # Query shifts with overtime with EXPLICIT JOIN condition
    shifts_query = db.session.query(
        Shift.date,
        Shift.start_time,
        Shift.end_time,
        User.section_id,
        User.unit_id,
        User.schedule_format
    ).join(User, Shift.employee_id == User.id).filter(  # EXPLICIT JOIN
        Shift.employee_id.in_(employee_ids),
        Shift.date.between(start_date, end_date),
        Shift.status == ShiftStatus.SCHEDULED,
        Shift.start_time.isnot(None),
        Shift.end_time.isnot(None)
    )
    
    # Add 4-level hierarchy fields if available
    if hasattr(User, 'department_id'):
        shifts_query = shifts_query.add_columns(User.department_id)
    if hasattr(User, 'division_id'):
        shifts_query = shifts_query.add_columns(User.division_id)
    
    shifts = shifts_query.all()
    
    # Calculate overtime hours for each shift
    overtime_data = []
    for shift in shifts:
        # Calculate shift duration
        start_datetime = datetime.combine(shift.date, shift.start_time)
        end_datetime = datetime.combine(shift.date, shift.end_time)
        
        # Handle shifts that cross midnight
        if end_datetime <= start_datetime:
            end_datetime += timedelta(days=1)
        
        total_hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        # Determine standard hours based on schedule format
        standard_hours = 9 if shift.schedule_format and 'nine' in shift.schedule_format.value else 8
        
        # Calculate overtime
        overtime_hours = max(0, total_hours - standard_hours)
        
        if overtime_hours > 0:
            overtime_data.append({
                'date': shift.date,
                'overtime_hours': overtime_hours,
                'section_id': shift.section_id,
                'unit_id': shift.unit_id,
                'department_id': getattr(shift, 'department_id', None),
                'division_id': getattr(shift, 'division_id', None)
            })
    
    # Process data based on view_type and group_by
    if view_type == 'daily':
        return process_daily_data(overtime_data, start_date, end_date, group_by, 'overtime')
    elif view_type == 'weekly':
        return process_weekly_data(overtime_data, start_date, end_date, group_by, 'overtime')
    elif view_type == 'monthly':
        return process_monthly_data(overtime_data, start_date, end_date, group_by, 'overtime')
    else:
        return []


def process_daily_data(data, start_date, end_date, group_by, data_type):
    """Process data for daily view"""
    result = []
    current_date = start_date
    
    while current_date <= end_date:
        day_data = [item for item in data if 
                   (item.extension_date if data_type == 'extension' else item['date']) == current_date]
        
        if group_by == 'total':
            total_hours = sum(
                item.extension_hours if data_type == 'extension' 
                else item['overtime_hours'] for item in day_data
            )
            result.append({
                'date': current_date.isoformat(),
                'label': current_date.strftime('%b %d'),
                'hours': float(total_hours),
                'group': 'Total'
            })
        else:
            # Group by organizational unit
            groups = defaultdict(float)
            for item in day_data:
                group_key = get_group_key(item, group_by, data_type)
                hours = item.extension_hours if data_type == 'extension' else item['overtime_hours']
                groups[group_key] += hours
            
            for group_name, hours in groups.items():
                result.append({
                    'date': current_date.isoformat(),
                    'label': current_date.strftime('%b %d'),
                    'hours': float(hours),
                    'group': group_name
                })
        
        current_date += timedelta(days=1)
    
    return result

def process_weekly_data(data, start_date, end_date, group_by, data_type):
    """Process data for weekly view"""
    result = []
    
    # Get start of first week (Monday)
    current_date = start_date - timedelta(days=start_date.weekday())
    
    while current_date <= end_date:
        week_end = current_date + timedelta(days=6)
        week_data = [item for item in data if 
                    current_date <= (item.extension_date if data_type == 'extension' else item['date']) <= week_end]
        
        week_label = f"{current_date.strftime('%b %d')} - {week_end.strftime('%b %d')}"
        
        if group_by == 'total':
            total_hours = sum(
                item.extension_hours if data_type == 'extension' 
                else item['overtime_hours'] for item in week_data
            )
            result.append({
                'date': current_date.isoformat(),
                'label': week_label,
                'hours': float(total_hours),
                'group': 'Total'
            })
        else:
            # Group by organizational unit
            groups = defaultdict(float)
            for item in week_data:
                group_key = get_group_key(item, group_by, data_type)
                hours = item.extension_hours if data_type == 'extension' else item['overtime_hours']
                groups[group_key] += hours
            
            for group_name, hours in groups.items():
                result.append({
                    'date': current_date.isoformat(),
                    'label': week_label,
                    'hours': float(hours),
                    'group': group_name
                })
        
        current_date += timedelta(days=7)
    
    return result

def process_monthly_data(data, start_date, end_date, group_by, data_type):
    """Process data for monthly view"""
    result = []
    
    # Get all months in the range
    current_date = start_date.replace(day=1)
    
    while current_date <= end_date:
        # Get last day of current month
        if current_date.month == 12:
            next_month = current_date.replace(year=current_date.year + 1, month=1)
        else:
            next_month = current_date.replace(month=current_date.month + 1)
        
        month_end = next_month - timedelta(days=1)
        
        month_data = [item for item in data if 
                     current_date <= (item.extension_date if data_type == 'extension' else item['date']) <= month_end]
        
        month_label = current_date.strftime('%b %Y')
        
        if group_by == 'total':
            total_hours = sum(
                item.extension_hours if data_type == 'extension' 
                else item['overtime_hours'] for item in month_data
            )
            result.append({
                'date': current_date.isoformat(),
                'label': month_label,
                'hours': float(total_hours),
                'group': 'Total'
            })
        else:
            # Group by organizational unit
            groups = defaultdict(float)
            for item in month_data:
                group_key = get_group_key(item, group_by, data_type)
                hours = item.extension_hours if data_type == 'extension' else item['overtime_hours']
                groups[group_key] += hours
            
            for group_name, hours in groups.items():
                result.append({
                    'date': current_date.isoformat(),
                    'label': month_label,
                    'hours': float(hours),
                    'group': group_name
                })
        
        current_date = next_month
    
    return result

def get_group_key(item, group_by, data_type):
    """Get the grouping key for an item"""
    if data_type == 'extension':
        if group_by == 'section':
            section = Section.query.get(item.section_id) if item.section_id else None
            return section.name if section else 'No Section'
        elif group_by == 'unit':
            unit = Unit.query.get(item.unit_id) if item.unit_id else None
            return unit.name if unit else 'No Unit'
        elif group_by == 'department' and hasattr(item, 'department_id'):
            from app.models import Department
            dept = Department.query.get(item.department_id) if item.department_id else None
            return dept.name if dept else 'No Department'
        elif group_by == 'division' and hasattr(item, 'division_id'):
            from app.models import Division
            div = Division.query.get(item.division_id) if item.division_id else None
            return div.name if div else 'No Division'
    else:  # overtime data
        if group_by == 'section':
            section = Section.query.get(item['section_id']) if item['section_id'] else None
            return section.name if section else 'No Section'
        elif group_by == 'unit':
            unit = Unit.query.get(item['unit_id']) if item['unit_id'] else None
            return unit.name if unit else 'No Unit'
        elif group_by == 'department' and item.get('department_id'):
            from app.models import Department
            dept = Department.query.get(item['department_id']) if item['department_id'] else None
            return dept.name if dept else 'No Department'
        elif group_by == 'division' and item.get('division_id'):
            from app.models import Division
            div = Division.query.get(item['division_id']) if item['division_id'] else None
            return div.name if div else 'No Division'
    
    return 'Unknown'

@bp.route('/api/organizational-units')
@login_required
def get_organizational_units():
    """Get available organizational units for grouping"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Available grouping options based on scope
    available_groups = ['total']
    
    if scope['type'] in ['admin', 'department', 'division']:
        available_groups.extend(['section', 'unit'])
        if hasattr(User, 'department_id'):
            available_groups.append('department')
        if hasattr(User, 'division_id'):
            available_groups.append('division')
    elif scope['type'] == 'section':
        available_groups.append('unit')
    
    return jsonify({
        'success': True,
        'available_groups': available_groups,
        'scope': scope
    })


# Add these additional routes to app/analytics/routes.py for data export functionality

@bp.route('/api/export-work-extension-data')
@login_required
def export_work_extension_data():
    """Export work extension data as CSV/Excel"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    format_type = request.args.get('format', 'csv')  # csv or excel
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    # Get detailed work extension data
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employee_ids = [emp.id for emp in employee_query.all()]
    
    if not employee_ids:
        return jsonify({'success': False, 'error': 'No employees found in scope'}), 400
    
    # Get work extensions with employee details using EXPLICIT JOIN
    extensions_query = db.session.query(
        WorkExtension.reference_code,
        WorkExtension.extension_date,
        WorkExtension.extension_hours,
        WorkExtension.reason,
        WorkExtension.status,
        WorkExtension.date_filed,
        WorkExtension.date_reviewed,
        User.first_name,
        User.last_name,
        User.personnel_number,
        Section.name.label('section_name'),
        Unit.name.label('unit_name')
    ).join(User, WorkExtension.employee_id == User.id)\
     .outerjoin(Section, User.section_id == Section.id)\
     .outerjoin(Unit, User.unit_id == Unit.id)\
     .filter(
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.extension_date.between(start_date, end_date),
        WorkExtension.status == WorkExtensionStatus.APPROVED
    ).order_by(WorkExtension.extension_date.desc())
    
    # Add computed column for full name
    extensions_query = extensions_query.add_columns(
        (func.concat(User.first_name, ' ', User.last_name)).label('employee_name')
    )
    
    # Add 4-level hierarchy if available
    if hasattr(User, 'department_id'):
        from app.models import Department
        extensions_query = extensions_query.outerjoin(Department, User.department_id == Department.id)\
                                         .add_columns(Department.name.label('department_name'))
    if hasattr(User, 'division_id'):
        from app.models import Division
        extensions_query = extensions_query.outerjoin(Division, User.division_id == Division.id)\
                                         .add_columns(Division.name.label('division_name'))
    
    extensions = extensions_query.all()
    
    if format_type == 'excel':
        return export_to_excel(extensions, 'work_extension', start_date, end_date, scope)
    else:
        return export_to_csv(extensions, 'work_extension', start_date, end_date, scope)

@bp.route('/api/export-overtime-data')
@login_required
def export_overtime_data():
    """Export overtime data as CSV/Excel"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    format_type = request.args.get('format', 'csv')
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    # Get overtime data from shifts
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employee_ids = [emp.id for emp in employee_query.all()]
    
    if not employee_ids:
        return jsonify({'success': False, 'error': 'No employees found in scope'}), 400
    
    # Get shifts with overtime using EXPLICIT JOIN
    shifts_query = db.session.query(
        Shift.date,
        Shift.start_time,
        Shift.end_time,
        User.first_name,
        User.last_name,
        User.personnel_number,
        User.schedule_format,
        Section.name.label('section_name'),
        Unit.name.label('unit_name')
    ).join(User, Shift.employee_id == User.id)\
     .outerjoin(Section, User.section_id == Section.id)\
     .outerjoin(Unit, User.unit_id == Unit.id)\
     .filter(
        Shift.employee_id.in_(employee_ids),
        Shift.date.between(start_date, end_date),
        Shift.status == ShiftStatus.SCHEDULED,
        Shift.start_time.isnot(None),
        Shift.end_time.isnot(None)
    ).order_by(Shift.date.desc())
    
    # Add computed column for full name
    shifts_query = shifts_query.add_columns(
        (func.concat(User.first_name, ' ', User.last_name)).label('employee_name')
    )
    
    # Add 4-level hierarchy if available
    if hasattr(User, 'department_id'):
        from app.models import Department
        shifts_query = shifts_query.outerjoin(Department, User.department_id == Department.id)\
                                  .add_columns(Department.name.label('department_name'))
    if hasattr(User, 'division_id'):
        from app.models import Division
        shifts_query = shifts_query.outerjoin(Division, User.division_id == Division.id)\
                                  .add_columns(Division.name.label('division_name'))
    
    shifts = shifts_query.all()
    
    # Calculate overtime for each shift
    overtime_data = []
    for shift in shifts:
        # Calculate shift duration
        start_datetime = datetime.combine(shift.date, shift.start_time)
        end_datetime = datetime.combine(shift.date, shift.end_time)
        
        # Handle shifts that cross midnight
        if end_datetime <= start_datetime:
            end_datetime += timedelta(days=1)
        
        total_hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        # Determine standard hours based on schedule format
        standard_hours = 9 if shift.schedule_format and 'nine' in shift.schedule_format.value else 8
        
        # Calculate overtime
        overtime_hours = max(0, total_hours - standard_hours)
        
        if overtime_hours > 0:
            overtime_record = {
                'date': shift.date,
                'employee_name': shift.employee_name,
                'personnel_number': shift.personnel_number or '',
                'section_name': shift.section_name or '',
                'unit_name': shift.unit_name or '',
                'shift_start': shift.start_time.strftime('%H:%M'),
                'shift_end': shift.end_time.strftime('%H:%M'),
                'total_hours': round(total_hours, 2),
                'standard_hours': standard_hours,
                'overtime_hours': round(overtime_hours, 2),
                'schedule_format': shift.schedule_format.value if shift.schedule_format else '8_hour_shift'
            }
            
            # Add 4-level hierarchy data if available
            if hasattr(shift, 'department_name'):
                overtime_record['department_name'] = shift.department_name or ''
            if hasattr(shift, 'division_name'):
                overtime_record['division_name'] = shift.division_name or ''
            
            overtime_data.append(overtime_record)
    
    if format_type == 'excel':
        return export_overtime_to_excel(overtime_data, start_date, end_date, scope)
    else:
        return export_overtime_to_csv(overtime_data, start_date, end_date, scope)

def export_to_csv(data, data_type, start_date, end_date, scope):
    """Export work extension data to CSV"""
    import io
    import csv
    from flask import make_response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    header = [
        'Reference Code', 'Employee Name', 'Personnel Number', 'Extension Date',
        'Extension Hours', 'Reason', 'Status', 'Date Filed', 'Date Reviewed',
        'Section', 'Unit'
    ]
    
    # Add 4-level hierarchy headers if available
    if hasattr(data[0], 'department_name') if data else False:
        header.insert(-2, 'Department')
    if hasattr(data[0], 'division_name') if data else False:
        header.insert(-2, 'Division')
    
    writer.writerow(header)
    
    # Write data
    for item in data:
        row = [
            item.reference_code,
            item.employee_name,
            item.personnel_number or '',
            item.extension_date.strftime('%Y-%m-%d'),
            item.extension_hours or 0,
            item.reason[:100] + '...' if len(item.reason) > 100 else item.reason,
            item.status.value,
            item.date_filed.strftime('%Y-%m-%d'),
            item.date_reviewed.strftime('%Y-%m-%d') if item.date_reviewed else '',
            item.section_name or '',
            item.unit_name or ''
        ]
        
        # Add 4-level hierarchy data if available
        if hasattr(item, 'department_name'):
            row.insert(-2, item.department_name or '')
        if hasattr(item, 'division_name'):
            row.insert(-2, item.division_name or '')
        
        writer.writerow(row)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=work_extensions_{scope["type"]}_{start_date}_{end_date}.csv'
    
    return response

def export_overtime_to_csv(data, start_date, end_date, scope):
    """Export overtime data to CSV"""
    import io
    import csv
    from flask import make_response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    header = [
        'Date', 'Employee Name', 'Personnel Number', 'Shift Start', 'Shift End',
        'Total Hours', 'Standard Hours', 'Overtime Hours', 'Schedule Format',
        'Section', 'Unit'
    ]
    
    # Add 4-level hierarchy headers if available
    if data and 'department_name' in data[0]:
        header.insert(-2, 'Department')
    if data and 'division_name' in data[0]:
        header.insert(-2, 'Division')
    
    writer.writerow(header)
    
    # Write data
    for item in data:
        row = [
            item['date'].strftime('%Y-%m-%d'),
            item['employee_name'],
            item['personnel_number'],
            item['shift_start'],
            item['shift_end'],
            item['total_hours'],
            item['standard_hours'],
            item['overtime_hours'],
            item['schedule_format'].replace('_', ' ').title(),
            item['section_name'],
            item['unit_name']
        ]
        
        # Add 4-level hierarchy data if available
        if 'department_name' in item:
            row.insert(-2, item['department_name'])
        if 'division_name' in item:
            row.insert(-2, item['division_name'])
        
        writer.writerow(row)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=overtime_data_{scope["type"]}_{start_date}_{end_date}.csv'
    
    return response

def export_to_excel(data, data_type, start_date, end_date, scope):
    """Export work extension data to Excel (requires openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import make_response
        import io
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Work Extensions"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Reference Code', 'Employee Name', 'Personnel Number', 'Extension Date',
            'Extension Hours', 'Reason', 'Status', 'Date Filed', 'Date Reviewed',
            'Section', 'Unit'
        ]
        
        # Add 4-level hierarchy headers if available
        if hasattr(data[0], 'department_name') if data else False:
            headers.insert(-2, 'Department')
        if hasattr(data[0], 'division_name') if data else False:
            headers.insert(-2, 'Division')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item.reference_code)
            ws.cell(row=row_num, column=2, value=item.employee_name)
            ws.cell(row=row_num, column=3, value=item.personnel_number or '')
            ws.cell(row=row_num, column=4, value=item.extension_date)
            ws.cell(row=row_num, column=5, value=item.extension_hours or 0)
            ws.cell(row=row_num, column=6, value=item.reason[:100] + '...' if len(item.reason) > 100 else item.reason)
            ws.cell(row=row_num, column=7, value=item.status.value)
            ws.cell(row=row_num, column=8, value=item.date_filed)
            ws.cell(row=row_num, column=9, value=item.date_reviewed if item.date_reviewed else '')
            
            col_offset = 0
            # Add 4-level hierarchy data if available
            if hasattr(item, 'department_name'):
                ws.cell(row=row_num, column=10, value=item.department_name or '')
                col_offset += 1
            if hasattr(item, 'division_name'):
                ws.cell(row=row_num, column=10 + col_offset, value=item.division_name or '')
                col_offset += 1
            
            ws.cell(row=row_num, column=10 + col_offset, value=item.section_name or '')
            ws.cell(row=row_num, column=11 + col_offset, value=item.unit_name or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=work_extensions_{scope["type"]}_{start_date}_{end_date}.xlsx'
        
        return response
        
    except ImportError:
        return jsonify({'success': False, 'error': 'Excel export requires openpyxl package'}), 500

def export_overtime_to_excel(data, start_date, end_date, scope):
    """Export overtime data to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import make_response
        import io
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overtime Data"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Date', 'Employee Name', 'Personnel Number', 'Shift Start', 'Shift End',
            'Total Hours', 'Standard Hours', 'Overtime Hours', 'Schedule Format',
            'Section', 'Unit'
        ]
        
        # Add 4-level hierarchy headers if available
        if data and 'department_name' in data[0]:
            headers.insert(-2, 'Department')
        if data and 'division_name' in data[0]:
            headers.insert(-2, 'Division')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item['date'])
            ws.cell(row=row_num, column=2, value=item['employee_name'])
            ws.cell(row=row_num, column=3, value=item['personnel_number'])
            ws.cell(row=row_num, column=4, value=item['shift_start'])
            ws.cell(row=row_num, column=5, value=item['shift_end'])
            ws.cell(row=row_num, column=6, value=item['total_hours'])
            ws.cell(row=row_num, column=7, value=item['standard_hours'])
            ws.cell(row=row_num, column=8, value=item['overtime_hours'])
            ws.cell(row=row_num, column=9, value=item['schedule_format'].replace('_', ' ').title())
            
            col_offset = 0
            # Add 4-level hierarchy data if available
            if 'department_name' in item:
                ws.cell(row=row_num, column=10, value=item['department_name'])
                col_offset += 1
            if 'division_name' in item:
                ws.cell(row=row_num, column=10 + col_offset, value=item['division_name'])
                col_offset += 1
            
            ws.cell(row=row_num, column=10 + col_offset, value=item['section_name'])
            ws.cell(row=row_num, column=11 + col_offset, value=item['unit_name'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=overtime_data_{scope["type"]}_{start_date}_{end_date}.xlsx'
        
        return response
        
    except ImportError:
        return jsonify({'success': False, 'error': 'Excel export requires openpyxl package'}), 500

@bp.route('/api/summary-report')
@login_required
def get_summary_report():
    """Get summary analytics report"""
    if not has_analytics_access():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    scope = get_user_scope()
    if not scope:
        return jsonify({'success': False, 'error': 'Unable to determine data scope'}), 403
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    # Get employees in scope
    employee_query = User.query.filter_by(is_active=True)
    employee_query = filter_employees_by_scope(employee_query, scope)
    employees = employee_query.all()
    employee_ids = [emp.id for emp in employees]
    
    if not employee_ids:
        return jsonify({'success': False, 'error': 'No employees found in scope'}), 400
    
    # Calculate summary statistics
    summary = {
        'period': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'days': (end_date - start_date).days + 1
        },
        'scope': scope,
        'employees': {
            'total': len(employees),
            'confidential': len([emp for emp in employees if emp.can_file_work_extension()])
        },
        'work_extensions': {},
        'overtime': {}
    }
    
    # Work extension statistics
    we_stats = db.session.query(
        func.count(WorkExtension.id).label('total_count'),
        func.sum(WorkExtension.extension_hours).label('total_hours'),
        func.avg(WorkExtension.extension_hours).label('avg_hours')
    ).filter(
        WorkExtension.employee_id.in_(employee_ids),
        WorkExtension.extension_date.between(start_date, end_date),
        WorkExtension.status == WorkExtensionStatus.APPROVED
    ).first()
    
    summary['work_extensions'] = {
        'total_extensions': we_stats.total_count or 0,
        'total_hours': float(we_stats.total_hours or 0),
        'average_hours': float(we_stats.avg_hours or 0),
        'average_per_employee': float(we_stats.total_hours or 0) / len(employees) if employees else 0
    }
    
    # Overtime statistics (calculated from shifts)
    shifts = db.session.query(
        Shift.date,
        Shift.start_time,
        Shift.end_time,
        User.schedule_format
    ).join(User).filter(
        Shift.employee_id.in_(employee_ids),
        Shift.date.between(start_date, end_date),
        Shift.status == ShiftStatus.SCHEDULED,
        Shift.start_time.isnot(None),
        Shift.end_time.isnot(None)
    ).all()
    
    total_overtime_hours = 0
    overtime_shifts = 0
    
    for shift in shifts:
        # Calculate shift duration
        start_datetime = datetime.combine(shift.date, shift.start_time)
        end_datetime = datetime.combine(shift.date, shift.end_time)
        
        if end_datetime <= start_datetime:
            end_datetime += timedelta(days=1)
        
        total_hours = (end_datetime - start_datetime).total_seconds() / 3600
        standard_hours = 9 if shift.schedule_format and 'nine' in shift.schedule_format.value else 8
        overtime_hours = max(0, total_hours - standard_hours)
        
        if overtime_hours > 0:
            total_overtime_hours += overtime_hours
            overtime_shifts += 1
    
    summary['overtime'] = {
        'total_shifts_with_overtime': overtime_shifts,
        'total_overtime_hours': round(total_overtime_hours, 2),
        'average_overtime_per_shift': round(total_overtime_hours / overtime_shifts, 2) if overtime_shifts > 0 else 0,
        'average_per_employee': round(total_overtime_hours / len(employees), 2) if employees else 0
    }
    
    return jsonify({
        'success': True,
        'summary': summary
    })